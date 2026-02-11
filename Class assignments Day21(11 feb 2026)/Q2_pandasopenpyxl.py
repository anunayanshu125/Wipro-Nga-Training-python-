import pandas as pd
from openpyxl import load_workbook

df = pd.read_excel("sales_data.xlsx", sheet_name="2025")

df["Total"] = df["Quantity"] * df["Price"]

df.to_excel("sales_summary.xlsx", index=False)

print("Pandas part completed - sales_summary.xlsx created")

wb = load_workbook("sales_data.xlsx")
ws = wb["2025"]

ws["D1"] = "Total"

for row in range(2, ws.max_row + 1):
    quantity = ws[f"B{row}"].value
    price = ws[f"C{row}"].value
    ws[f"D{row}"] = quantity * price

wb.save("sales_summary_openpyxl.xlsx")

print("OpenPyXL part completed - sales_summary_openpyxl.xlsx created")